"""
Сервис для автозаполнения документов
"""
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
from flask import current_app

# ВРЕМЕННОЕ РЕШЕНИЕ: Добавляем путь к системным библиотекам
miniconda_path = r'C:\Users\Boot\miniconda3\lib\site-packages'
if os.path.exists(miniconda_path) and miniconda_path not in sys.path:
    sys.path.insert(0, miniconda_path)
    print(f"🔧 Добавлен путь к системным библиотекам: {miniconda_path}")

# Простая проверка импорта без сложных альтернативных методов
HAS_DOCXTPL = False
HAS_OPENPYXL = False
DocxTemplate = None
openpyxl = None

try:
    from docxtpl import DocxTemplate
    HAS_DOCXTPL = True
    print("✅ docxtpl успешно импортирован")
except ImportError as e:
    print(f"❌ docxtpl не найден: {e}")
    print("💡 Попробуйте: pip install docxtpl")

try:
    import openpyxl
    HAS_OPENPYXL = True
    print("✅ openpyxl успешно импортирован")
except ImportError as e:
    print(f"❌ openpyxl не найден: {e}")
    print("💡 Попробуйте: pip install openpyxl")

print(f"📊 Итоговый статус: docxtpl={HAS_DOCXTPL}, openpyxl={HAS_OPENPYXL}")

from app.models import Client


class DocumentService:
    """Сервис для работы с документами клиентов"""
    
    def __init__(self):
        self.docs_folder: Optional[Path] = None
        self.work_folder: Optional[Path] = None
    
    def _init_folders(self) -> None:
        """Инициализация папок"""
        if self.docs_folder is None:
            self.docs_folder = Path(current_app.config.get('DOCS_FOLDER', './docs'))
        if self.work_folder is None:
            self.work_folder = Path(current_app.config.get('WORK_FOLDER', './work_files'))
    
    def get_service_documents(self, service_id: int) -> List[str]:
        """Получить список документов для конкретной услуги"""
        document_mapping = {
            0: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],  # Акт обследования
            1: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],  # Выдел
            2: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],  # Вынос
            3: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],  # Образование
            4: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docх'],  # Объединение
            5: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],  # Перераспределение
            6: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],  # Раздел
            7: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx', 
                'Декларация.xlsm', 'Акт к договору подряда.docx'],  # Технический план
            8: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],  # Уточнение
        }
        return document_mapping.get(service_id, [])
    
    def copy_documents_to_folder(self, folder_path: str, service_id: int, force_copy: bool = False) -> Dict[str, Any]:
        """
        Копирование шаблонов документов в указанную папку
        
        Args:
            folder_path: Путь к папке клиента
            service_id: ID услуги для определения нужных документов
            force_copy: Принудительное копирование даже если файлы существуют
            
        Returns:
            Dict с результатами копирования
        """
        self._init_folders()
        
        if not os.path.exists(folder_path):
            return {
                'success': False,
                'errors': [f'Папка не существует: {folder_path}'],
                'copied_files': []
            }
        
        if self.docs_folder is None or not self.docs_folder.exists():
            return {
                'success': False,
                'errors': [f'Папка шаблонов не найдена: {self.docs_folder}'],
                'copied_files': []
            }
        
        client_folder = Path(folder_path)
        required_docs = self.get_service_documents(service_id)
        
        copied_files = []
        errors = []
        
        for doc_name in required_docs:
            source_path = self.docs_folder / doc_name
            dest_path = client_folder / doc_name
            
            # Копируем если файл не существует или принудительное копирование
            if source_path.exists() and (not dest_path.exists() or force_copy):
                try:
                    shutil.copy2(source_path, dest_path)
                    copied_files.append(doc_name)
                    print(f"✅ Скопирован: {doc_name}")
                except Exception as e:
                    errors.append(f"Ошибка копирования {doc_name}: {str(e)}")
                    print(f"❌ Ошибка копирования {doc_name}: {e}")
            elif not source_path.exists():
                errors.append(f"Шаблон не найден: {doc_name}")
                print(f"⚠️ Шаблон не найден: {doc_name}")
        
        return {
            'success': len(copied_files) > 0 or len(errors) == 0,
            'copied_files': copied_files,
            'errors': errors
        }

    def copy_documents_to_client_folder(self, client: Client, force_copy: bool = False) -> Dict[str, Any]:
        """
        Копирование шаблонов документов в папку клиента
        
        Args:
            client: Объект клиента
            force_copy: Принудительное копирование даже если файлы существуют
            
        Returns:
            Dict с результатами копирования
        """
        self._init_folders()
        
        if not client.path_folder or not os.path.exists(client.path_folder):
            return {
                'success': False,
                'errors': ['Папка клиента не существует'],
                'copied_files': []
            }
        
        if self.docs_folder is None or not self.docs_folder.exists():
            return {
                'success': False,
                'errors': ['Папка шаблонов не найдена'],
                'copied_files': []
            }
        
        client_folder = Path(client.path_folder)
        required_docs = self.get_service_documents(client.service)
        
        copied_files = []
        errors = []
        
        for doc_name in required_docs:
            source_path = self.docs_folder / doc_name
            dest_path = client_folder / doc_name
            
            # Копируем если файл не существует или принудительное копирование
            if source_path.exists() and (not dest_path.exists() or force_copy):
                try:
                    shutil.copy2(source_path, dest_path)
                    copied_files.append(doc_name)
                    print(f"✅ Скопирован: {doc_name}")
                except Exception as e:
                    errors.append(f"Ошибка копирования {doc_name}: {str(e)}")
                    print(f"❌ Ошибка копирования {doc_name}: {e}")
            elif not source_path.exists():
                errors.append(f"Шаблон не найден: {doc_name}")
                print(f"⚠️ Шаблон не найден: {doc_name}")
        
        return {
            'success': len(copied_files) > 0 or len(errors) == 0,
            'copied_files': copied_files,
            'errors': errors
        }
    
    def ensure_documents_in_folder(self, client: Client, required_docs: List[str]) -> Dict[str, Any]:
        """
        Убеждаемся что нужные документы есть в папке клиента
        Копируем только те, которых нет
        
        Args:
            client: Объект клиента
            required_docs: Список нужных документов
            
        Returns:
            Dict с результатами
        """
        self._init_folders()
        
        if not client.path_folder or not os.path.exists(client.path_folder):
            return {
                'success': False,
                'errors': ['Папка клиента не существует'],
                'copied_files': []
            }
        
        client_folder = Path(client.path_folder)
        copied_files = []
        errors = []
        
        # Маппинг ключей документов к файлам
        doc_mapping = {
            'approval': 'Согласие.docx',
            'contract': 'Договор подряда.docx',
            'contract_agreement': 'Акт к договору подряда.docx',
            'declaration': 'Декларация.xlsm',
            'receipt': 'Квитанция.xlsx'
        }
        
        for doc_key in required_docs:
            if doc_key in doc_mapping:
                doc_filename = doc_mapping[doc_key]
                source_path = self.docs_folder / doc_filename
                dest_path = client_folder / doc_filename
                
                # Копируем только если файла нет в папке клиента
                if not dest_path.exists():
                    if source_path.exists():
                        try:
                            shutil.copy2(source_path, dest_path)
                            copied_files.append(doc_filename)
                            print(f"✅ Скопирован нужный документ: {doc_filename}")
                        except Exception as e:
                            errors.append(f"Ошибка копирования {doc_filename}: {str(e)}")
                    else:
                        errors.append(f"Шаблон не найден: {doc_filename}")
        
        return {
            'success': len(errors) == 0,
            'copied_files': copied_files,
            'errors': errors
        }
    
    def get_context_data(self, client: Client) -> Dict[str, str]:
        """Получить данные для подстановки в шаблоны"""
        # Создаем дату договора
        if client.work_info and client.work_info.date_status:
            contract_date = client.work_info.date_status
        else:
            contract_date = datetime.now().strftime('%d.%m.%Y')
        
        # Парсим дату для номера договора
        try:
            date_parts = contract_date.split('.')
            day = date_parts[0]
            month = date_parts[1]
        except (IndexError, AttributeError):
            day = datetime.now().strftime('%d')
            month = datetime.now().strftime('%m')
        
        context = {
            # Номер и дата договора
            'number_contact_agreement': f"{client.id}-{day}/{month}",
            'date_contract_agreement': contract_date,
            
            # Личные данные
            'sur_name': client.sur_name or '',
            'name': client.name or '',
            'middle_name': client.middle_name or '',
            'telefone': client.telefone or '',
            
            # Адрес
            'address': client.address.address if client.address else '',
            
            # Паспортные данные
            'series_pass': client.doc_info.series_pass if client.doc_info else '',
            'date_pass': client.doc_info.date_pass if client.doc_info else '',
            'info_pass': client.doc_info.info_pass if client.doc_info else '',
            'snils': client.doc_info.snils if client.doc_info else '',
            
            # Дополнительные данные
            'date_burn': client.doc_fill_info.date_birthday if client.doc_fill_info else '',
            'place_residence': client.doc_fill_info.place_residence if client.doc_fill_info else '',
            'service_extended': client.doc_fill_info.extend_work_info if client.doc_fill_info else client.service_name,
        }
        
        return context
    
    def fill_word_document(self, template_path: Path, output_path: Path, context: Dict[str, str]) -> bool:
        """Заполнение Word документа"""
        if not HAS_DOCXTPL or DocxTemplate is None:
            print("❌ docxtpl недоступен")
            return False
            
        try:
            doc = DocxTemplate(str(template_path))
            doc.render(context)
            doc.save(str(output_path))
            print(f"✅ Word документ создан: {output_path}")
            return True
        except Exception as e:
            print(f"❌ Ошибка создания Word документа {template_path}: {e}")
            return False
    
    def fill_excel_document(self, template_path: Path, output_path: Path, context: Dict[str, str], doc_type: str) -> bool:
        """Заполнение Excel документа"""
        if not HAS_OPENPYXL or openpyxl is None:
            print("❌ openpyxl недоступен")
            return False
            
        try:
            wb = openpyxl.load_workbook(str(template_path))
            ws = wb.active
            
            if ws is None:
                print(f"❌ Активный лист не найден в {template_path}")
                wb.close()
                return False
            
            if doc_type == 'Декларация.xlsm':
                # Заполняем декларацию
                ws['Q144'] = context['sur_name']
                ws['Q145'] = context['name']
                ws['Q146'] = context['middle_name']
                ws['AB147'] = context['snils']
                ws['S148'] = context['series_pass']
                cell_d149_value = f"{context['date_pass']} {context['info_pass']}" if context['date_pass'] and context['info_pass'] else ""
                ws['D149'] = cell_d149_value
                
            elif doc_type == 'Квитанция.xlsx':
                # Заполняем квитанцию
                full_name = f"{context['sur_name']} {context['name']} {context['middle_name']}".strip()
                ws['C21'] = full_name
                ws['C22'] = context['service_extended']
                ws['B24'] = f"По адресу: {context['address']}"
                ws['F14'] = context['number_contact_agreement']
            
            wb.save(str(output_path))
            wb.close()
            print(f"✅ Excel документ создан: {output_path}")
            return True
        except Exception as e:
            print(f"❌ Ошибка создания Excel документа {template_path}: {e}")
            return False
    
    def fill_client_documents(self, client_id: int, selected_docs: Optional[List[str]] = None, auto_copy: bool = True) -> Dict[str, Any]:
        """
        Основная функция заполнения документов клиента
        
        Args:
            client_id: ID клиента
            selected_docs: Список выбранных документов для заполнения
            auto_copy: Автоматически копировать отсутствующие документы
        """
        from app import db
        
        client = Client.query.get(client_id)
        if not client:
            raise ValueError("Клиент не найден")
        
        if not client.path_folder or not os.path.exists(client.path_folder):
            raise ValueError("Папка клиента не существует")
        
        # Проверяем доступность библиотек
        if not HAS_DOCXTPL and not HAS_OPENPYXL:
            return {
                'success': False,
                'errors': ['Библиотеки для работы с документами недоступны. Установите: pip install docxtpl openpyxl'],
                'filled_documents': [],
                'copied_documents': []
            }
        
        # Если включено автокопирование, сначала копируем нужные документы
        copied_documents = []
        copy_errors = []
        
        if auto_copy and selected_docs:
            copy_result = self.ensure_documents_in_folder(client, selected_docs)
            copied_documents = copy_result['copied_files']
            copy_errors = copy_result['errors']
            
            if copy_errors:
                print(f"⚠️ Предупреждения при копировании: {copy_errors}")
        
        # Проверяем обязательные поля
        missing_fields = []
        if not client.doc_info or not client.doc_info.series_pass:
            missing_fields.append("Серия/номер паспорта")
        if not client.doc_info or not client.doc_info.info_pass:
            missing_fields.append("Кем выдан паспорт")
        if not client.doc_info or not client.doc_info.snils:
            missing_fields.append("СНИЛС")
        if not client.doc_fill_info or not client.doc_fill_info.extend_work_info:
            missing_fields.append("Подробное наименование работ")
        if not client.doc_fill_info or not client.doc_fill_info.place_residence:
            missing_fields.append("Место проживания")
        
        if missing_fields:
            raise ValueError(f"Не заполнены обязательные поля: {', '.join(missing_fields)}")
        
        client_folder = Path(client.path_folder)
        context = self.get_context_data(client)
        
        filled_documents: List[str] = []
        errors: List[str] = copy_errors.copy()  # Начинаем с ошибок копирования
        
        # Определяем какие документы нужно заполнить
        docs_to_fill: Dict[str, str] = {}
        
        # Если переданы конкретные документы из формы
        if selected_docs:
            doc_mapping = {
                'approval': ('Согласие.docx', 'Согласие.docx'),
                'contract': ('Договор подряда.docx', 'Договор подряда.docx'),
                'contract_agreement': ('Акт к договору подряда.docx', 'Акт к договору подряда.docx'),
                'declaration': ('Декларация.xlsm', 'Декларация.xlsm'),
                'receipt': ('Квитанция.xlsx', 'Квитанция.xlsx')
            }
            
            for doc_key in selected_docs:
                if doc_key in doc_mapping:
                    template_name, output_name = doc_mapping[doc_key]
                    docs_to_fill[template_name] = output_name
        else:
            # Используем настройки из базы данных
            if client.doc_fill_info:
                if client.doc_fill_info.approval:
                    docs_to_fill['Согласие.docx'] = 'Согласие.docx'
                if client.doc_fill_info.contract:
                    docs_to_fill['Договор подряда.docx'] = 'Договор подряда.docx'
                if client.doc_fill_info.contract_agreement:
                    docs_to_fill['Акт к договору подряда.docx'] = 'Акт к договору подряда.docx'
                if client.doc_fill_info.declaration:
                    docs_to_fill['Декларация.xlsm'] = 'Декларация.xlsm'
                if client.doc_fill_info.receipt:
                    docs_to_fill['Квитанция.xlsx'] = 'Квитанция.xlsx'
        
        # Заполняем документы
        for template_name, output_name in docs_to_fill.items():
            template_path = client_folder / template_name
            output_path = client_folder / output_name
            
            if not template_path.exists():
                errors.append(f"Шаблон {template_name} не найден в папке клиента")
                continue
            
            success = False
            if template_name.endswith('.docx'):
                if HAS_DOCXTPL:
                    success = self.fill_word_document(template_path, output_path, context)
                else:
                    errors.append(f"Пропущен {template_name} - docxtpl недоступен")
            elif template_name.endswith('.xlsx'):
                if HAS_OPENPYXL:
                    success = self.fill_excel_document(template_path, output_path, context, template_name)
                else:
                    errors.append(f"Пропущен {template_name} - openpyxl недоступен")
            
            if success:
                filled_documents.append(output_name)
                print(f"✅ Документ заполнен: {output_name}")
            else:
                if template_name.endswith('.docx') and HAS_DOCXTPL:
                    errors.append(f"Ошибка при заполнении {template_name}")
                elif template_name.endswith('.xlsx') and HAS_OPENPYXL:
                    errors.append(f"Ошибка при заполнении {template_name}")
        
        return {
            'success': len(filled_documents) > 0,
            'filled_documents': filled_documents,
            'copied_documents': copied_documents,
            'errors': errors,
            'client_folder': str(client_folder)
        }


# Создаем экземпляр сервиса
document_service = DocumentService()